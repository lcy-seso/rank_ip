#coding=utf8
import os
import sys
import gzip
import pdb
import math
import numpy as np
from openpyxl import Workbook

import paddle.v2 as paddle
import reader
from network import half_ranknet


def norm_score(x):
    return float(1. / (1 + math.exp(-x)) * 100)


def get_file_name(file_path):
    return os.path.splitext(os.path.split(file_path)[1])[0]


def infer(model_path, test_dir, raw_data_dir, batch_size):
    assert os.path.exists(model_path), "The trained model does not exist."
    feature_dim = 7

    parameters = paddle.parameters.Parameters.from_tar(gzip.open(model_path))
    inferer = paddle.inference.Inference(
        output_layer=half_ranknet("infer", feature_dim), parameters=parameters)

    res_file = Workbook()
    res_sheets = []
    files = os.listdir(test_dir)
    files.sort()
    sheet_titles = open("sheet_title_05_06.txt").readlines()

    all_samples = []
    for i, f_name in enumerate(files):
        test_file = os.path.join(test_dir, f_name)
        raw_file = os.path.join(raw_data_dir, f_name)
        print("predicting : %s\nraw_file: %s" % (test_file, raw_file))

        fraw = open(raw_file, "r")
        titles = fraw.readline().strip().split("\t")

        fn = get_file_name(raw_file)
        res_sheet = res_file.create_sheet(str(i))
        res_sheet.title = sheet_titles[i].strip().decode("utf-8")

        idx = 1
        # write header for each sheet.
        res_sheet["A%d" % idx] = u"分值"
        for t_id, title in enumerate(titles):
            res_sheet["%s%d" % (chr(ord("B") + t_id), idx)] = title

        test_batch = []
        scores = []
        raw_inputs = []
        for idx, item in enumerate(reader.test_reader(test_file)()):
            test_batch.append([item])
            raw_inputs.append(fraw.readline().strip().split("\t"))

            if len(test_batch) == batch_size:
                scores += inferer.infer(
                    input=test_batch, field=["value"]).tolist()
                test_batch = []

        if len(test_batch):
            scores += inferer.infer(input=test_batch, field=["value"]).tolist()
            test_batch = []
        predictions = zip(scores, raw_inputs)

        idx = 2
        for p in predictions:
            score = norm_score(p[0][0])
            all_samples.append([score, p[1], sheet_titles[i], titles])

            res_sheet["A%d" % (idx)] = "%.3f" % (score)
            for j, context in enumerate(p[1]):
                res_sheet["%s%d" % (chr(ord("B") + j), idx)] = context
            idx += 1

    res_sheet = res_file.create_sheet(str(len(files)))
    res_sheet.title = u"所有数据"
    idx = 1
    res_sheet["A%d" % idx] = u"分值"
    res_sheet["B%d" % idx] = u"来源"

    count = 2
    book_name = set()
    for i, sample in enumerate(all_samples):
        if sample[1][0] in book_name: continue

        res_sheet["A%d" % (count)] = "%.3f" % sample[0]
        res_sheet["B%d" % (count)] = "%s" % sample[2]

        for j, context in enumerate(sample[1]):
            res_sheet["%s%d" % (chr(ord("C") + j),
                                count)] = "%s:  %s" % (sample[-1][j], context)

        book_name.add(sample[1][0])
        count += 1

    res_file.save("2018_05_06_scores.xlsx")


if __name__ == "__main__":
    model_path = "models_05_05/ranknet_params_00008.tar.gz"
    test_dir = "processed/final"
    raw_data_dir = ("/Users/ying/Documents/codes/rank_ip/preprocess/"
                    "data/2018_05_05/raw_data")

    paddle.init(use_gpu=False, trainer_count=1)
    infer(model_path, test_dir, raw_data_dir, batch_size=16)
