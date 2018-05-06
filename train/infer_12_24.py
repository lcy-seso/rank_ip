#coding=utf8
import os
import sys
import gzip
import pdb
import math
import numpy as np
from openpyxl import Workbook

import paddle.v2 as paddle
import reader
from network import half_ranknet_12_24


def norm_score(x):
    return float(1. / (1 + math.exp(-x)) * 100)


def get_file_name(file_path):
    return os.path.splitext(os.path.split(file_path)[1])[0]


def infer(model_path,
          test_dir,
          raw_data_dir,
          batch_size,
          data_info=("/Users/ying/Documents/codes/"
                     "rank_ip/preprocess/data/12_24/data_info.txt")):
    assert os.path.exists(model_path), "The trained model does not exist."
    feature_dim = 7

    parameters = paddle.parameters.Parameters.from_tar(gzip.open(model_path))
    inferer = paddle.inference.Inference(
        output_layer=half_ranknet_12_24("infer", feature_dim),
        parameters=parameters)

    res_file = Workbook()
    res_sheets = []
    files = os.listdir(test_dir)
    files.sort()
    sheet_titles = open("sheet_title.txt").readlines()

    all_samples = []
    for i, f_name in enumerate(files):
        test_file = os.path.join(test_dir, f_name)
        raw_file = os.path.join(raw_data_dir, f_name)
        fn = get_file_name(raw_file)

        raw_sample = {}
        with open(raw_file, "r") as fin:
            for line in fin:
                line_split = line.strip().split("\t")
                raw_sample[line_split[0]] = line.strip()

        res_sheet = res_file.create_sheet(str(i))
        res_sheet.title = sheet_titles[i].strip().decode("utf-8")

        idx = 1
        res_sheet["A%d" % idx] = u"分值"

        if i > 5:
            title = open(("/Users/ying/Documents/codes/rank_ip/"
                          "preprocess/data/12_26/data_info.txt"),
                         "r").readlines()
        else:
            titles = open(data_info, "r").readlines()

        for t, title in enumerate(titles):
            res_sheet["%s%d" % (chr(ord("B") + t), idx)] = title.split(",")[0]

        test_batch = []
        info = []
        scores = []
        for idx, item in enumerate(reader.test_reader(test_file)()):
            test_batch.append([item[0]])
            info.append(item[1])

            if len(test_batch) == batch_size:
                scores += inferer.infer(
                    input=test_batch, field=["value"]).tolist()
                test_batch = []

        if len(test_batch):
            scores += inferer.infer(input=test_batch, field=["value"]).tolist()
            test_batch = []
        predictions = zip(scores, info)

        idx = 2
        for p in predictions:
            score = norm_score(p[0][0])
            res_sheet["A%d" % (idx)] = "%.2f" % (score)

            for j, context in enumerate(raw_sample[p[1]].split("\t")):
                res_sheet["%s%d" % (chr(ord("B") + j), idx)] = context
            all_samples.append(
                [sheet_titles[i].strip(), score, raw_sample[p[1]]])
            idx += 1

    res_sheet = res_file.create_sheet(str(len(files)))
    res_sheet.title = u"所有数据"
    idx = 1
    res_sheet["A%d" % idx] = u"分值"
    res_sheet["B%d" % idx] = u"来源"

    titles = open(data_info, "r").readlines()
    for t, title in enumerate(titles):
        res_sheet["%s%d" % (chr(ord("C") + t), idx)] = title.split(",")[0]

    sample_set = set()
    count = 0
    for i, sample in enumerate(all_samples):
        title = sample[2].split("\t")[0]
        if title in sample_set: continue
        res_sheet["A%d" % (count + 2)] = "%.2f" % sample[1]
        sample_set.add(title)
        res_sheet["B%d" % (count + 2)] = sample[0].decode("utf-8")
        for j, context in enumerate(sample[2].split("\t")):
            res_sheet["%s%d" % (chr(ord("C") + j), count + 2)] = context
        count += 1

    res_file.save("12_26_scores.xlsx")


if __name__ == '__main__':
    # model_path = "models_12_24/ranknet_params_000040.tar.gz"
    model_path = "models_12_26/ranknet_params_000050.tar.gz"
    test_dir = "test_data"
    raw_data_dir = ("/Users/ying/Documents/codes/rank_ip/preprocess/"
                    "data/12_26/data_for_report")

    paddle.init(use_gpu=False, trainer_count=1)
    infer(model_path, test_dir, raw_data_dir, 1)
